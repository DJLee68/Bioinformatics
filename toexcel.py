# -*- coding: utf-8 -*-
"""
Created on Tue Oct 30 18:14:03 2018

@author: leedongjae
"""

# import modules
import pandas as pd
import numpy as np
from openpyxl import load_workbook 
from openpyxl import Workbook


idx = 2
row_num = 1
excel_sheetname = ""
def excel_init():
    wb = Workbook()
    wb.save(filename='./test.xlsx')
    
def sheet_init(file_name):    
    global excel_sheetname
    excel_sheetname = file_name.split('/')[-1]
    
    wb = load_workbook(filename = './test.xlsx')
    ws = wb.create_sheet(title = excel_sheetname)
    
    # index_format(index) & columns_format(columns)정의
    ws.cell(row=1, column=2).value = '전처리기'
    ws.cell(row=1, column=3).value = '특징 선택'
    ws.cell(row=1, column=4).value = 'PCA 수'
    ws.cell(row=1, column=5).value = '분류기'
    ws.cell(row=1, column=6).value = 'KNN 개수'
    ws.cell(row=1, column=7).value = 'Accuracy'
    ws.cell(row=1, column=8).value = 'Precision'
    ws.cell(row=1, column=9).value = 'Recall'
    ws.cell(row=1, column=10).value = 'F1-score'
    ws.cell(row=1, column=11).value = 'Support'
    ws.cell(row=1, column=12).value = 'Confusion-matrix'

    wb.save(filename='./test.xlsx')

def set_value(preprocessor_name, fselector_name, pca_num, classifier_name, knn_k, result_value, precision, recall, fbeta_score, support, conf_mat):
    global idx
    global row_num
    global excel_sheetname
    excel_filename = './test.xlsx'
    wb = load_workbook(filename = excel_filename)
    
    # get the worksheet

    
    ws = wb.get_sheet_by_name(excel_sheetname)
    ws.cell(row=idx, column=1).value = row_num
    ws.cell(row=idx, column=2).value = preprocessor_name
    ws.cell(row=idx, column=3).value = fselector_name
    ws.cell(row=idx, column=4).value = pca_num
    ws.cell(row=idx, column=5).value = classifier_name
    
    if knn_k is not None:
        ws.cell(row=idx, column=6).value = knn_k
    
    ws.cell(row=idx, column=7).value = result_value
    
    for i in range(precision.size):
        ws.cell(row=idx+i, column=8).value = precision[i]
        
    for i in range(recall.size):
        ws.cell(row=idx+i, column=9).value = recall[i]
        
    for i in range(fbeta_score.size):
        ws.cell(row=idx+i, column=10).value = fbeta_score[i]
    
    for i in range(support.size):
        ws.cell(row=idx+i, column=11).value = support[i]
        
    for i in range(conf_mat[0].size):
        for j in range(conf_mat[0].size):
            ws.cell(row=idx+i, column=12+j).value = conf_mat[i][j]
            
    row_num+=1
    idx+=precision.size
    
    wb.save(filename=excel_filename)
